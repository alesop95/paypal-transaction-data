import os
import logging
from typing import List, Dict, Optional
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import Config

class ExcelClient:
    def __init__(self):
        self.excel_file = Config.EXCEL_FILE_PATH
        self.sheet_name = Config.EXCEL_SHEET_NAME
        
        self.logger = logging.getLogger(__name__)
        
        # Assicurati che la directory esista
        Path(self.excel_file).parent.mkdir(parents=True, exist_ok=True)
    
    def _get_sheet_headers(self) -> List[str]:
        """Define the headers for the PayPal transactions sheet"""
        return [
            'Transaction ID',
            'PayPal Reference ID', 
            'Order ID',
            'Transaction Date',
            'Updated Date',
            'Status',
            'Event Code',
            'Subject',
            'Gross Amount',
            'Currency',
            'Fee Amount',
            'Net Amount',
            'Payer Name',
            'Payer Email',
            'Payer Account ID',
            'Payer Country',
            'Invoice ID',
            'Custom Field',
            'Protection Eligibility',
            'Extracted At',
            'API Mode'
        ]
    
    def create_or_update_sheet(self):
        """Create the Excel file if it doesn't exist or update headers"""
        try:
            if os.path.exists(self.excel_file):
                # Carica il workbook esistente
                workbook = load_workbook(self.excel_file)
                
                # Controlla se il sheet esiste
                if self.sheet_name in workbook.sheetnames:
                    worksheet = workbook[self.sheet_name]
                else:
                    worksheet = workbook.create_sheet(self.sheet_name)
            else:
                # Crea nuovo workbook
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = self.sheet_name
            
            # Controlla se ci sono già headers
            if worksheet.max_row == 1 and worksheet['A1'].value is None:
                # Aggiungi headers
                headers = self._get_sheet_headers()
                for col, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=1, column=col, value=header)
                    # Formatta header (grassetto e sfondo grigio)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                
                # Auto-adjust column widths
                for col in range(1, len(headers) + 1):
                    column_letter = get_column_letter(col)
                    worksheet.column_dimensions[column_letter].width = 15
                
                # Salva il file
                workbook.save(self.excel_file)
                self.logger.info(f"Created Excel file with headers: {self.excel_file}")
            else:
                self.logger.info(f"Excel file already exists with data: {self.excel_file}")
            
        except Exception as e:
            self.logger.error(f"Error creating/updating Excel file: {e}")
            raise
    
    def get_existing_transaction_ids(self) -> set:
        """Get all existing transaction IDs from the Excel file to avoid duplicates"""
        try:
            if not os.path.exists(self.excel_file):
                return set()
            
            workbook = load_workbook(self.excel_file)
            
            if self.sheet_name not in workbook.sheetnames:
                return set()
            
            worksheet = workbook[self.sheet_name]
            
            transaction_ids = set()
            
            # Leggi dalla colonna A (Transaction ID), saltando l'header
            for row in range(2, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row, column=1).value
                if cell_value:
                    transaction_ids.add(str(cell_value).strip())
            
            self.logger.info(f"Found {len(transaction_ids)} existing transactions in Excel file")
            return transaction_ids
            
        except Exception as e:
            self.logger.error(f"Error reading existing transactions: {e}")
            return set()
    
    def append_transactions(self, transactions: List[Dict]) -> int:
        """
        Append new transactions to the Excel file
        
        Args:
            transactions: List of transaction dictionaries with fiscal data
            
        Returns:
            Number of transactions actually added
        """
        if not transactions:
            self.logger.info("No transactions to append")
            return 0
        
        # Get existing transaction IDs to avoid duplicates
        existing_ids = self.get_existing_transaction_ids()
        
        # Filter out existing transactions
        new_transactions = [
            t for t in transactions 
            if t.get('transaction_id') and t['transaction_id'] not in existing_ids
        ]
        
        if not new_transactions:
            self.logger.info("No new transactions to add (all already exist)")
            return 0
        
        try:
            # Carica o crea il workbook
            if os.path.exists(self.excel_file):
                workbook = load_workbook(self.excel_file)
                worksheet = workbook[self.sheet_name]
            else:
                self.create_or_update_sheet()
                workbook = load_workbook(self.excel_file)
                worksheet = workbook[self.sheet_name]
            
            # Trova la prossima riga vuota
            next_row = worksheet.max_row + 1
            
            # Aggiungi le transazioni
            for transaction in new_transactions:
                row_data = [
                    transaction.get('transaction_id', ''),
                    transaction.get('paypal_reference_id', ''),
                    transaction.get('order_id', ''),
                    transaction.get('transaction_date', ''),
                    transaction.get('transaction_updated_date', ''),
                    transaction.get('transaction_status', ''),
                    transaction.get('transaction_event_code', ''),
                    transaction.get('transaction_subject', ''),
                    transaction.get('gross_amount', ''),
                    transaction.get('currency_code', ''),
                    transaction.get('fee_amount', ''),
                    transaction.get('net_amount', ''),
                    transaction.get('payer_name', ''),
                    transaction.get('payer_email', ''),
                    transaction.get('payer_account_id', ''),
                    transaction.get('payer_country_code', ''),
                    transaction.get('invoice_id', ''),
                    transaction.get('custom_field', ''),
                    transaction.get('protection_eligibility', ''),
                    transaction.get('extracted_at', ''),
                    transaction.get('api_mode', '')
                ]
                
                for col, value in enumerate(row_data, 1):
                    worksheet.cell(row=next_row, column=col, value=value)
                
                next_row += 1
            
            # Salva il file
            workbook.save(self.excel_file)
            
            added_count = len(new_transactions)
            self.logger.info(f"Successfully added {added_count} new transactions to Excel file")
            
            return added_count
            
        except Exception as e:
            self.logger.error(f"Error appending transactions to Excel file: {e}")
            raise
    
    def update_transaction_status(self, transaction_id: str, new_status: str) -> bool:
        """Update the status of a specific transaction"""
        try:
            if not os.path.exists(self.excel_file):
                self.logger.warning(f"Excel file not found: {self.excel_file}")
                return False
            
            workbook = load_workbook(self.excel_file)
            worksheet = workbook[self.sheet_name]
            
            # Cerca la transazione nella colonna A
            for row in range(2, worksheet.max_row + 1):
                if worksheet.cell(row=row, column=1).value == transaction_id:
                    # Aggiorna lo status (colonna F)
                    worksheet.cell(row=row, column=6, value=new_status)
                    workbook.save(self.excel_file)
                    
                    self.logger.info(f"Updated status for transaction {transaction_id} to {new_status}")
                    return True
            
            self.logger.warning(f"Transaction {transaction_id} not found in Excel file")
            return False
            
        except Exception as e:
            self.logger.error(f"Error updating transaction status: {e}")
            return False
    
    def get_sheet_summary(self) -> Dict:
        """Get summary information about the Excel file"""
        try:
            if not os.path.exists(self.excel_file):
                return {
                    'total_transactions': 0,
                    'last_update': None,
                    'currencies': [],
                    'status_counts': {},
                    'file_path': self.excel_file
                }
            
            workbook = load_workbook(self.excel_file)
            worksheet = workbook[self.sheet_name]
            
            if worksheet.max_row <= 1:  # Solo headers o vuoto
                return {
                    'total_transactions': 0,
                    'last_update': None,
                    'currencies': [],
                    'status_counts': {},
                    'file_path': self.excel_file
                }
            
            # Analizza i dati
            total_count = worksheet.max_row - 1  # Escludi header
            
            # Raccogli valute e status
            currencies = set()
            statuses = {}
            last_update = None
            
            for row in range(2, worksheet.max_row + 1):
                # Currency (colonna J)
                currency = worksheet.cell(row=row, column=10).value
                if currency:
                    currencies.add(currency)
                
                # Status (colonna F)
                status = worksheet.cell(row=row, column=6).value
                if status:
                    statuses[status] = statuses.get(status, 0) + 1
                
                # Last update (colonna T)
                extracted_at = worksheet.cell(row=row, column=20).value
                if extracted_at:
                    last_update = extracted_at
            
            return {
                'total_transactions': total_count,
                'last_update': last_update,
                'currencies': list(currencies),
                'status_counts': statuses,
                'file_path': self.excel_file
            }
            
        except Exception as e:
            self.logger.error(f"Error getting Excel file summary: {e}")
            return {
                'total_transactions': 0,
                'last_update': None,
                'currencies': [],
                'status_counts': {},
                'file_path': self.excel_file,
                'error': str(e)
            }
